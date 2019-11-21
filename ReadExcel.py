import xlrd as xlrd
from openpyxl import *


class ReadExcel(object):
    def __init__(self):
        # self.file = file
        # self.wb = load_workbook(self.file)
        # sheets = self.wb.get_sheet_names()
        # self.sheet = sheets[0]
        # self.ws = self.wb[self.sheet]
        pass

    # 获取表格的总行数和总列数
    def get_rows_columns(self, file):
        """ 获取表格的总行数和总列数

        Examples:
        | ${getData} | get_rows_columns | C:\\Users\\Administrator\\Desktop\\123.xlsx |

        Return:
        | 元组 | (rows,colums) |

        """
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def get_cell_value(self, row, column, file):
        """ 获取某个单元格的值

        Examples:
        | ${getOneData} | get_cell_value | row | column | C:\\Users\\Administrator\\Desktop\\123.xlsx |

        Return:
        |  cellvalue |

        """
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]
        cellvalue = self.ws.cell(row=row, column=column).value
        return cellvalue

    # 获取某列的所有值
    def get_col_values(self, column, file):
        """ 获取某列的所有值

        Examples:
        | ${getColValues} | get_col_values | column | C:\\Users\\Administrator\\Desktop\\123.xlsx |

        Return:
        | 列表 | [colums...] |

        """
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]
        rows = self.ws.max_row
        columndata = []
        for i in range(1, rows + 1):
            cellvalue = self.ws.cell(row=i, column=column).value
            columndata.append(cellvalue)
        return columndata

    # 获取某行所有值
    def get_row_values(self, row, file):
        """ 获取某行所有值

        Examples:
        | ${getRowValues} | get_row_values | row | C:\\Users\\Administrator\\Desktop\\123.xlsx |

        Return:
        | 列表 | [rowdatas...] |

        """
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]
        columns = self.ws.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = self.ws.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    # 获取EXCEL所有数据以列表字典形式返回
    def get_all_values(self, file, sheet_name):
        """ 获取EXCEL所有数据以列表字典形式返回

        Examples:
        | ${getAllValues} | get_all_values | C:\\Users\\Administrator\\Desktop\\123.xlsx | sheet_name |

        Return:
        | 列表 | [{"name":"张三","age":23},{"name":"李四","age":23}] |

        """
        bk = xlrd.open_workbook(file)
        sh = bk.sheet_by_name(sheet_name)
        row_num = sh.nrows
        data_list = []
        for i in range(1, row_num):
            row_data = sh.row_values(i)
            data = {}
            for index, key in enumerate(sh.row_values(0)):
                data[key] = row_data[index]
            data_list.append(data)
        return data_list

    # # 设置某个单元格的值
    #     # def set_cell_value(self, row, colunm, cellvalue, file):
    #     #     self.file = file
    #     #     self.wb = load_workbook(self.file)
    #     #     sheets = self.wb.get_sheet_names()
    #     #     self.sheet = sheets[0]
    #     #     self.ws = self.wb[self.sheet]
    #     #     try:
    #     #         self.ws.cell(row=row, column=colunm).value = cellvalue
    #     #         self.wb.save(self.file)
    #     #     except:
    #     #         self.ws.cell(row=row, column=colunm).value = "writefail"
    #     #         self.wb.save(self.file)
